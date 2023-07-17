import sys
import os
import datetime
import numpy as np
import pandas as pd
import rstr
import random
import subprocess
import shutil

from PyQt6.QtCore import Qt, QDateTime
from PyQt6.QtWidgets import QApplication
from PyQt6.QtWidgets import QLabel
from PyQt6.QtWidgets import (
    QWidget,
    QFileDialog,
    QMainWindow,
    QMenuBar,
    QMenu,
    QGridLayout,
    QLineEdit,
    QDateEdit,
    QPushButton,
    QMessageBox,
)

from openpyxl import load_workbook


class Window(QMainWindow):
    def __init__(self, parent=None):
        """Initializer for the entire GUI"""
        super().__init__(parent)
        self.recovery_path = os.getcwd() + os.path.sep + "recovery"
        os.makedirs(self.recovery_path) if not os.path.exists(self.recovery_path) else None  # creates recovery folder
        self.site = "Hamburg"
        self.alloc = [1 if i % 2 == 0 else -1 for i in range(6)]  # allocates the first subjects randomly
        self.file_label = QLabel()
        self.file_label.setText('Keine Datei bisher ausgewählt')

        self.setWindowTitle("Randomisierung ParkProReakt - {}".format(self.site))
        self.resize(400, 200)
        self._createMenuBar()
        self._check_for_config()
        self._createForm()

    def _createMenuBar(self):
        """creates a menu which served to select the Excel file of interest"""
        menuBar = QMenuBar()
        self.setMenuBar(menuBar)

        # Creating menus using a QMenu object
        fileMenu = QMenu(" &Datei", self)
        fileMenu.addAction("Exceltabelle auswählen", self._selectFile)
        menuBar.addMenu(fileMenu)

    def _createForm(self):
        # Widget und Layout
        self.form_widget = QWidget(parent=self)
        self.grid_layout = QGridLayout()
        self.form_widget.setLayout(self.grid_layout)

        # Formular Label und Felder
        surname_label = QLabel(self.form_widget)
        surname_label.setText("Nachname: ")
        self.surname_value = QLineEdit()

        name_label = QLabel(self.form_widget)
        name_label.setText("Vorname: ")
        self.name_value = QLineEdit()

        birthday_label = QLabel(self.form_widget)
        birthday_label.setText("Geburtsdatum: ")
        self.birthday_value = QDateEdit(calendarPopup=True)
        self.birthday_value.setDateTime(QDateTime.currentDateTime())
        self.birthday_value.setDisplayFormat("dd.MM.yyyy")

        bdi_label = QLabel(self.form_widget)
        bdi_label.setText("Beck's Depressions Inventar: ")
        self.bdi_value = QLineEdit()

        hy_value = QLabel(self.form_widget)
        hy_value.setText("Hoehn und Yahr Stadium: ")
        self.hy_value = QLineEdit()

        pdq8_value = QLabel(self.form_widget)
        pdq8_value.setText("Parkinson's Disease Questionnaire (PDQ-8): ")
        self.pdq8_value = QLineEdit()

        self.file_label = QLabel()
        self.file_label.setText("Ausgewählte Datei: " + str(self.excel_file))

        # Buttons
        self.clearButton = QPushButton()
        self.clearButton.setText("Felder löschen")

        self.addButton = QPushButton()
        self.addButton.setText("Patient hinzufügen")

        self.openXLS = QPushButton()
        self.openXLS.setText("Bisherige Liste öffnen")

        self.addButton.clicked.connect(self.addPatient)
        self.clearButton.clicked.connect(self.clearForm)
        self.openXLS.clicked.connect(self.openXLS_file)

        # Layout fpr gesamtes Formular
        self.grid_layout.addWidget(self.file_label, 0, 0, 1, 2)

        self.grid_layout.addWidget(name_label, 1, 0)
        self.grid_layout.addWidget(self.name_value, 1, 1)
        self.grid_layout.addWidget(surname_label, 2, 0)
        self.grid_layout.addWidget(self.surname_value, 2, 1)
        self.grid_layout.addWidget(birthday_label, 3, 0)
        self.grid_layout.addWidget(self.birthday_value, 3, 1)
        self.grid_layout.addWidget(bdi_label, 4, 0)
        self.grid_layout.addWidget(self.bdi_value, 4, 1)
        self.grid_layout.addWidget(hy_value, 5, 0)
        self.grid_layout.addWidget(self.hy_value, 5, 1)
        self.grid_layout.addWidget(pdq8_value, 6, 0)
        self.grid_layout.addWidget(self.pdq8_value, 6, 1)
        self.grid_layout.addWidget(self.clearButton, 7, 0)
        self.grid_layout.addWidget(self.addButton, 7, 1)
        self.grid_layout.addWidget(self.openXLS, 8, 1)

        self.setCentralWidget(self.form_widget)
        self.grid_layout.setAlignment(
            Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter
        )

    def create_backup(self):
        """this is an additional backup feature in case the template is chosen, so that no data is lost"""

        workbook = load_workbook(os.path.join(os.getcwd(), "randomisation{}.xlsx".format(self.site)))
        current_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        workbook.save(os.path.join(self.recovery_path, current_date + 'randomisation{}.xlsx'.format(self.site)))

    def _selectFile(self):
        self.excel_file = QFileDialog.getOpenFileName(self, 'Exceltabelle auswählen',
                                                      filter='Excel files (*.xls *.xlsx)', directory=os.getcwd())[0]
        if self.excel_file == os.path.join(os.getcwd(), 'template.xlsx'):
            try:
                self.create_backup()
            except FileNotFoundError:
                print('No file to backup, creating new xlsx-file')

            shutil.copyfile(os.path.join(os.getcwd(), 'template.xlsx'), "randomisation{}.xlsx".format(self.site))
            self.excel_file = os.path.join(os.getcwd(), "randomisation{}.xlsx".format(self.site))

        with open(os.getcwd() + os.path.sep + 'config.ini', 'w') as config_file:
            config_file.write(self.excel_file)
        self.file_label.setText('Ausgewählte Datei: ' + str(self.excel_file))

    def _check_for_config(self):
        if "config.ini" in os.listdir(os.getcwd()):
            with open(os.getcwd() + os.path.sep + 'config.ini', 'r') as config_file:
                self.excel_file = config_file.read()
        else:
            self._selectFile()

    def clearForm(self):
        """deletes all entries in case something has been entered in the wrong way"""
        for item in self.form_widget.findChildren((QLineEdit, QDateEdit)):
            if isinstance(item, QLineEdit):
                item.setText("")
            elif isinstance(item, QDateEdit):
                item.setDateTime(QDateTime.currentDateTime())

    def openXLS_file(self):
        """opens the stored data in the standard program"""
        subprocess.call(['open', self.excel_file])

    @staticmethod
    def get_maximum_rows(sheet):
        """helper function to obtain the last entry in the xlsx-file. Source:
        https://stackoverflow.com/questions/46569496/openpyxl-max-row-and-max-column-wrongly-reports-a-larger-figure"""

        for i in range(1, 20000):
            if sheet.cell(row=i, column=2).value is None:
                max_row = i
                break
        return max_row

    def addPatient(self):
        """adds the data that was entered in the form into an Excel file that is stored locally. Furthermore,
        every time the xlsx-file is read a copy is saved to prevent data loss"""

        workbook, sheet, df = self.load_excelfile(self)
        # load_workbook(self.excel_file)
        current_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        workbook.save(os.path.join(self.recovery_path, current_date+'.recovery.xlsx'))  # recovery save
        current_patients = self.get_maximum_rows(sheet)

        birthday = self.birthday_value.dateTime().toPyDateTime()
        age = (datetime.datetime.now() - birthday) / datetime.timedelta(days=365.25)
        dialog_response = self._showDialog(current_patients-1)

        if current_patients > 7:
            bt, fim_total = self.initialise_randomisation(df)

        if dialog_response == QMessageBox.StandardButton.Ok:
            sheet["A" + str(current_patients)] = self.name_value.text()
            sheet["B" + str(current_patients)] = self.surname_value.text()
            sheet["C" + str(current_patients)] = self.create_pseudonym(8)
            sheet[
                "D" + str(current_patients)
                ] = self.birthday_value.dateTime().toString("dd.MM.yyyy")
            sheet["E" + str(current_patients)] = age
            sheet["F" + str(current_patients)] = float(self.bdi_value.text())
            sheet["G" + str(current_patients)] = float(self.hy_value.text())
            sheet["H" + str(current_patients)] = float(self.pdq8_value.text())
            if current_patients < 8:
                sheet["I" + str(current_patients)] = self.alloc[current_patients-2]
            else:
                df_temp = pd.DataFrame({'intercept': [1],
                                        'bdi': [float(self.bdi_value.text())],
                                        'hy': [float(self.hy_value.text())],
                                        'pdq8': [float(self.pdq8_value.text())]})
                alloc_temp = self.patient_allocation(df_patient=df_temp, bt=bt, fim_total=fim_total)
                sheet["I" + str(current_patients)] = alloc_temp

            workbook.save(self.excel_file)
            self.clearForm()
        else:
            return

    @staticmethod
    def load_excelfile(self):
        """loads patients and their details of randomisation"""
        workbook = load_workbook(self.excel_file)
        sheet = workbook.active
        data = sheet.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)

        return workbook, sheet, df

    def _showDialog(self, number_of_patients: float):
        """Generates a message box to make sure that data is entered correctly into the database"""
        msg = QMessageBox()
        msg.setWindowTitle("Patient:innen hinzufügen")
        msg.setText(
            f"Bist Du sicher, dass Du {self.name_value.text()} {self.surname_value.text()} als {number_of_patients}. "
            f"Patient:in in {self.site} zur ParkProReakt hinzufügen möchtest?")
        msg.setStandardButtons(QMessageBox.StandardButton.Ok | QMessageBox.StandardButton.Cancel)

        return msg.exec()

    @staticmethod
    def create_pseudonym(size_array: int) -> str:
        """generates pseudonym of letters, numbers and special characters; ';' and ','
        are omitted to avoid confusion in 'csv-files' """
        re_expression = f'[a-zA-Z0-9_!#%$§]{{{size_array}}}'
        return rstr.xeger(re_expression)

    @staticmethod
    def get_irule(site: str) -> str:
        """according to the publication/simulation, there are different strategies for the centres especially because
        of distinct sample sizes, for details cf.: 10.1016/j.conctc.2023.101140"""

        irule_mapping = {
            "Hamburg": "deterministic",
            "Marburg": "ACA"
        }
        return irule_mapping.get(site, "default_value")

    def initialise_randomisation(self, current_df, nv=3, npat_first=6):
        """uses the algorithm developed, which aims at optimal allocation of subjects according to some rules defined
        for details cf.: 10.1016/j.conctc.2023.101140"""

        current_df = current_df.iloc[:, [5, 6, 7]]  # 0 and 1 are the column indexes

        # Normalize current dataframe and add column of 'ones' in front
        current_df = (current_df - np.min(current_df, axis=0)) / \
                     (np.max(current_df, axis=0) - np.min(current_df, axis=0))
        current_df = current_df.assign(intercept=1)
        current_df = current_df.reindex(columns=['intercept', 'bdi', 'hy', 'pdq8'])
        nv = nv + 1

        # Compute FIM for each treatment
        fim_groups = np.zeros((nv, nv, 2))
        for i in range(npat_first):
            fim_indiv = np.outer(current_df.iloc[i, :nv], current_df.iloc[i, :nv])
            if self.alloc[i] == -1:
                fim_groups[:, :, 0] += fim_indiv
            else:
                fim_groups[:, :, 1] += fim_indiv

        # Count the number of patients already allocated to each treatment
        nalloc = [self.alloc.count(-1), self.alloc.count(1)]
        fim_total = sum(nalloc[i] / npat_first * fim_groups[:, :, i] for i in range(2))

        # Variables accumulating covariates
        bt = np.zeros(nv)

        for i in range(npat_first):
            bt += current_df.iloc[i, 0:nv] * self.alloc[i]

        return bt, fim_total

    def patient_allocation(self, df_patient: pd.DataFrame, bt: int, fim_total: int, nv: int = 4):
        """This is the core of the allocation/randomisation algorithm proposed in our manuscript. According to two
        distinct rules the subject of interest is allocated into either of the arms. This complex procedure aims at
        balancing both groups according to some predictors that were deemed inportant for quality of life in PD and
        that should be balanced after inclusion of all participants"""

        irule = self.get_irule(self.site)
        inv_fim = np.linalg.inv(fim_total)
        d1 = df_patient.iloc[0, :nv] @ inv_fim @ bt
        rule = 0.5 - d1 / (1.0 + d1 ** 2)
        ran_val = random.random()  # randomization

        if irule == 'deterministic':
            alloc = 1 if ran_val <= 0.5 else -1
        else:  # ACA rule
            alloc = 1 if ran_val < rule else -1

        return alloc


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = Window()
    win.setFocusPolicy(Qt.FocusPolicy.ClickFocus)
    win.setFocus()
    win.show()
    sys.exit(app.exec())
