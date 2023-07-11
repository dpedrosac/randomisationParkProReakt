from faker import Faker
import os
import pandas as pd
import rstr
import random
from openpyxl import load_workbook
import datetime
from dateutil import parser
from dateutil.relativedelta import relativedelta
import numpy as np
import matplotlib.pyplot as plt


fake = Faker()
alloc = [1 if i % 2 == 0 else -1 for i in range(6)]  # allocates the first subjects randomly
excel_file = '/media/storage/randomisationParkProReakt/test/test1.xlsx'

@staticmethod
def load_excelfile():
    """loads patients and their details of randomisation"""
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    data = sheet.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    return workbook, sheet, df

@staticmethod
def initialise_randomisation(current_df, alloc, nv=3, npat_first=6):
    """uses the algorithm developed, which aims at optimal allocation of subjects according to some rules defined
    for details cf.: 10.1016/j.conctc.2023.101140"""

    current_df = current_df.iloc[:, [5, 6, 7]]  # 0 and 1 are the column indexes

    # Normalize current dataframe and add column of 'ones' in front
    current_df = (current_df - np.min(current_df, axis=0)) / \
                 (np.max(current_df, axis=0) - np.min(current_df, axis=0))
    current_df = current_df.assign(intercept=1)
    current_df = current_df.reindex(columns=['intercept', 'bdi', 'hy', 'pdq8'])
    nv = nv + 1

    # Compute FIM for each treatment
    fim_groups = np.zeros((nv, nv, 2))
    for i in range(npat_first):
        fim_indiv = np.outer(current_df.iloc[i, :nv], current_df.iloc[i, :nv])
        if alloc[i] == -1:
            fim_groups[:, :, 0] += fim_indiv
        else:
            fim_groups[:, :, 1] += fim_indiv

    # Count the number of patients already allocated to each treatment
    nalloc = [alloc.count(-1), alloc.count(1)]
    fim_total = sum(nalloc[i] / npat_first * fim_groups[:, :, i] for i in range(2))

    # Variables accumulating covariates
    bt = np.zeros(nv)

    for i in range(npat_first):
        bt += current_df.iloc[i, 0:nv] * alloc[i]

    return bt, fim_total

@staticmethod
def create_pseudonym(size_array: int) -> str:
    """generates pseudonym of letters, numbers and special characters; ';' and ','
    are omitted to avoid confusion in 'csv-files' """
    re_expression = f'[a-zA-Z0-9_!#%$§]{{{size_array}}}'
    return rstr.xeger(re_expression)

@staticmethod
def get_irule(site: str) -> str:
    """according to the publication/simulation, there are different strategies for the centres especially because
    of distinct sample sizes, for details cf.: 10.1016/j.conctc.2023.101140"""

    irule_mapping = {
        "Hamburg": "deterministic",
        "Marburg": "ACA"
    }
    return irule_mapping.get(site, "default_value")


@staticmethod
def patient_allocation(df_patient: pd.DataFrame, bt: int, fim_total: int, npat_first: int = 6, nv: int = 4):
    """This is the core of the allocation/randomisation algorithm proposed in our manuscript. According to two
    distinct rules the subject of interest is allocated into either of the arms. This complex procedure aims at
    balancing both groups according to some predictors that were deemed inportant for quality of life in PD and
    that should be balanced after inclusion of all participants"""

    site = 'Marburg'
    irule = get_irule(site)
    inv_fim = np.linalg.inv(fim_total)
    d1 = df_patient.iloc[0, :nv] @ inv_fim @ bt
    rule = 0.5 - d1 / (1.0 + d1 ** 2)
    ran_val = random.random()  # randomization

    if irule == 'deterministic':
        alloc = 1 if ran_val <= 0.5 else -1
    else:  # ACA rule
        alloc = 1 if ran_val < rule else -1

    return alloc


@staticmethod
def addPatient(firstname, lastname, bdi, hy, pdq8, birthdate, alloc):
    """adds the data that was entered in the form into an Excel file that is stored locally. Furthermore,
    every time the Excel file is read a copy is saved to prevent data loss"""

    excel_file = '/media/storage/randomisationParkProReakt/test/test1.xlsx'
    workbook, sheet, df = load_excelfile()
    current_patients = sheet.max_row - 1
    age = (datetime.datetime.now() - birthdate) / datetime.timedelta(days=365.25)
    if current_patients > 5:
        bt, fim_total = initialise_randomisation(df, alloc, nv=3, npat_first=6)

    sheet["A" + str(current_patients + 2)] = firstname
    sheet["B" + str(current_patients + 2)] = lastname
    sheet["C" + str(current_patients + 2)] = create_pseudonym(8)
    sheet[
        "D" + str(current_patients + 2)
        ] = birthdate.strftime('%d-%m-%Y')
    sheet["E" + str(current_patients + 2)] = age
    sheet["F" + str(current_patients + 2)] = bdi
    sheet["G" + str(current_patients + 2)] = hy
    sheet["H" + str(current_patients + 2)] = pdq8
    if current_patients+2 < 8:
        sheet["I" + str(current_patients + 2)] = alloc[current_patients]
    else:
        df_temp = pd.DataFrame({'intercept': [1],
                                'bdi': [bdi],
                                'hy': [hy],
                                'pdq8': [pdq8]})
        alloc_temp = patient_allocation(df_patient=df_temp, bt=bt, fim_total=fim_total)
        sheet["I" + str(current_patients + 2)] = alloc_temp

    workbook.save(excel_file)


for i in range(132):
    firstname = fake.first_name()
    lastname = fake.last_name()
    bdi = random.randint(1, 36)
    hy = random.randint(1 ,4)
    pdq8 = random.random()*100
    start_date = parser.parse('1942-01-01')
    end_date = parser.parse('1991-01-01')
    random_date = start_date + relativedelta(days=random.randint(0, (end_date - start_date).days))
    addPatient(firstname, lastname, bdi, hy, pdq8, random_date, alloc)


excel_file = '/media/storage/randomisationParkProReakt/test/test1.xlsx'

# Read the Excel file into a DataFrame
df = pd.read_excel(excel_file)

# Extract the necessary columns
data = df.iloc[:, 4:7]  # Assuming columns 5 to 7 are indexed as 4 to 6
category = df.iloc[:, 8].tolist()  # Assuming column 8 is indexed as 7

# Create boxplots
data.boxplot(by=category, figsize=(8, 6))

# Set plot labels and title
plt.xlabel('Category')
plt.ylabel('Values')
plt.title('Boxplots of Columns 5 to 7 by Category')

# Display the plot
plt.show()